from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from collections import defaultdict
import simplejson


def parse_locations():
    wb = load_workbook('locations/hacknu-dev-data.xlsx')
    data = {}
    for worksheet in wb:
        if worksheet.title not in data:
            data[worksheet.title] = []
        for row in worksheet.iter_rows(min_row=2, min_col=1, values_only=True):
            if row[0] is not None:
                d = {
                    'latitude':row[0], 
                    'longitude':row[1],
                    'altitude':row[2],
                    'identifier':row[3],
                    'timestamp':row[4],
                    'floor_label':row[5],
                    'horizontal_acc':row[6],
                    'vertical_acc':row[7],
                    'confidence':row[8],
                    'activity':row[9]
                }
                data[worksheet.title].append(d)
    return simplejson.dumps(data)
